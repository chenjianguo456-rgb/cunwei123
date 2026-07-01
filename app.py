#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
村委资料共享管理系统
功能：乡村振兴、党员信息、村民信息、养老信息、耕地信息、资料台账
特性：响应式设计、身份证校验、性别/年龄自动计算、Excel导入导出、后台管理
管理员账户：admin / admin123456
"""

import os, json, re, datetime, io, base64
from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_file, Response
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import Workbook, load_workbook

app = Flask(__name__, template_folder='templates', static_folder='static')

# 生产环境配置
is_production = os.environ.get('RENDER') or os.environ.get('PYTHONANYWHERE_DOMAIN')
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'cunwei-2024-secret-key-8f3a')

# 数据库URI处理：Render的PostgreSQL地址以postgres://开头，SQLAlchemy需要postgresql://
database_url = os.environ.get('DATABASE_URL', 'sqlite:///cunwei.db')
if database_url.startswith('postgres://'):
    database_url = database_url.replace('postgres://', 'postgresql://', 1)
elif database_url.startswith('postgresql://') and '+psycopg2' not in database_url:
    # 确保使用psycopg2驱动
    database_url = database_url.replace('postgresql://', 'postgresql+psycopg2://', 1)
app.config['SQLALCHEMY_DATABASE_URI'] = database_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['JSON_AS_ASCII'] = False
# 生产环境数据库连接池配置
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'pool_pre_ping': True,
    'pool_recycle': 300,
}

# 生产环境使用绝对路径
if is_production:
    base_dir = os.environ.get('HOME', '/app')
    app.config['UPLOAD_FOLDER'] = os.path.join(base_dir, 'cunwei_uploads')
else:
    app.config['UPLOAD_FOLDER'] = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'uploads')

app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB上传限制
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

db = SQLAlchemy(app)

# ===================== 数据库会话管理 =====================

@app.teardown_appcontext
def shutdown_session(exception=None):
    """确保每个请求结束后正确回滚或关闭数据库会话"""
    try:
        if exception:
            db.session.rollback()
    finally:
        db.session.remove()

@app.errorhandler(Exception)
def handle_exception(error):
    """全局异常处理：回滚会话并返回友好错误信息"""
    db.session.rollback()
    app.logger.error(f"Unhandled error: {error}", exc_info=True)
    # 如果是AJAX请求，返回JSON
    if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({'error': '服务器内部错误，请稍后重试'}), 500
    return '服务器内部错误，请稍后重试', 500

# ===================== 工具函数 =====================

def validate_id_card(id_card):
    """校验中国大陆18位身份证号码"""
    if not id_card or len(id_card) != 18:
        return False
    pattern = r'^\d{17}[\dXx]$'
    if not re.match(pattern, id_card):
        return False
    weights = [7,9,10,5,8,4,2,1,6,3,7,9,10,5,8,4,2]
    check_codes = ['1','0','X','9','8','7','6','5','4','3','2']
    try:
        sum_val = sum(int(id_card[i]) * weights[i] for i in range(17))
        return check_codes[sum_val % 11].upper() == id_card[17].upper()
    except Exception:
        return False

def extract_gender(id_card):
    """从身份证号提取性别：奇数为男，偶数为女"""
    if not id_card or len(id_card) < 17:
        return ''
    try:
        return '男' if int(id_card[16]) % 2 == 1 else '女'
    except Exception:
        return ''

def extract_age(id_card):
    """从身份证号提取年龄"""
    if not id_card or len(id_card) < 14:
        return ''
    try:
        year = int(id_card[6:10])
        month = int(id_card[10:12])
        day = int(id_card[12:14])
        today = datetime.date.today()
        age = today.year - year
        if (today.month, today.day) < (month, day):
            age -= 1
        return str(age)
    except Exception:
        return ''

def auto_fill_value(column, row_data):
    """根据列类型自动填充值"""
    col_type = column.get('type', 'text')
    source = column.get('source', '')
    if col_type == 'auto' and source == 'id_card_gender':
        id_card = row_data.get('id_card', '') or row_data.get('身份证号', '') or row_data.get('身份证号码', '')
        return extract_gender(id_card)
    if col_type == 'auto' and source == 'id_card_age':
        id_card = row_data.get('id_card', '') or row_data.get('身份证号', '') or row_data.get('身份证号码', '')
        return extract_age(id_card)
    return row_data.get(column['key'], '')

# ===================== 数据库模型 =====================

class User(db.Model):
    __tablename__ = 'users'
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(200), nullable=False)
    is_admin = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.datetime.now)

class Section(db.Model):
    __tablename__ = 'sections'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    icon = db.Column(db.String(50), default='📁')
    icon_image = db.Column(db.String(200), default='')  # 上传的图标图片文件名
    sort_order = db.Column(db.Integer, default=0)
    # 增色板块：颜色主题
    bg_color = db.Column(db.String(20), default='')      # 背景色（如 bg-green-50 或 hex）
    text_color = db.Column(db.String(20), default='')    # 文字颜色
    accent_color = db.Column(db.String(20), default='')  # 强调色/边框色
    description = db.Column(db.String(200), default='')   # 分区描述文字
    created_at = db.Column(db.DateTime, default=datetime.datetime.now)

class SiteConfig(db.Model):
    __tablename__ = 'site_config'
    id = db.Column(db.Integer, primary_key=True)
    key = db.Column(db.String(100), unique=True, nullable=False)
    value = db.Column(db.Text, default='')

class Template(db.Model):
    __tablename__ = 'templates'
    id = db.Column(db.Integer, primary_key=True)
    section_id = db.Column(db.Integer, db.ForeignKey('sections.id'), nullable=False)
    name = db.Column(db.String(100), nullable=False)
    columns = db.Column(db.Text, default='[]')  # JSON格式列定义
    created_at = db.Column(db.DateTime, default=datetime.datetime.now)

class SubSection(db.Model):
    __tablename__ = 'sub_sections'
    id = db.Column(db.Integer, primary_key=True)
    section_id = db.Column(db.Integer, db.ForeignKey('sections.id'), nullable=False)
    name = db.Column(db.String(100), nullable=False)
    icon = db.Column(db.String(50), default='📄')
    icon_image = db.Column(db.String(200), default='')
    description = db.Column(db.String(200), default='')
    sort_order = db.Column(db.Integer, default=0)
    columns = db.Column(db.Text, default='[]')  # JSON 模板列定义
    created_at = db.Column(db.DateTime, default=datetime.datetime.now)
    
    section = db.relationship('Section', backref='sub_sections')

class Record(db.Model):
    __tablename__ = 'records'
    id = db.Column(db.Integer, primary_key=True)
    section_id = db.Column(db.Integer, db.ForeignKey('sections.id'), nullable=False)
    template_id = db.Column(db.Integer, db.ForeignKey('templates.id'), nullable=False)
    subsection_id = db.Column(db.Integer, db.ForeignKey('sub_sections.id'), nullable=True)
    data = db.Column(db.Text, default='{}')  # JSON格式行数据
    created_at = db.Column(db.DateTime, default=datetime.datetime.now)

class DeletedRecord(db.Model):
    __tablename__ = 'deleted_records'
    id = db.Column(db.Integer, primary_key=True)
    original_id = db.Column(db.Integer, nullable=False)
    section_id = db.Column(db.Integer, db.ForeignKey('sections.id'), nullable=False)
    template_id = db.Column(db.Integer, db.ForeignKey('templates.id'), nullable=False)
    subsection_id = db.Column(db.Integer, db.ForeignKey('sub_sections.id'), nullable=True)
    data = db.Column(db.Text, default='{}')
    deleted_by = db.Column(db.String(80), nullable=False)
    deleted_at = db.Column(db.DateTime, default=datetime.datetime.now)

class UploadedImage(db.Model):
    """存储上传的图片（base64），替代文件系统——云部署不会丢失"""
    __tablename__ = 'uploaded_images'
    id = db.Column(db.Integer, primary_key=True)
    filename = db.Column(db.String(200), unique=True, nullable=False)
    mime_type = db.Column(db.String(50), default='image/png')
    data_b64 = db.Column(db.Text, default='')  # base64编码的图片数据
    created_at = db.Column(db.DateTime, default=datetime.datetime.now)

# ===================== 初始化数据 =====================

def get_config(key, default=''):
    """读取全局配置"""
    try:
        cfg = SiteConfig.query.filter_by(key=key).first()
        return cfg.value if cfg else default
    except Exception:
        return default

def set_config(key, value):
    """写入全局配置"""
    cfg = SiteConfig.query.filter_by(key=key).first()
    if cfg:
        cfg.value = value
    else:
        cfg = SiteConfig(key=key, value=value)
        db.session.add(cfg)
    safe_commit()

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'png', 'jpg', 'jpeg', 'gif', 'webp', 'bmp'}

def safe_commit(label=''):
    """安全提交：捕获异常、回滚、记录日志，并抛出带详细信息的错误"""
    try:
        safe_commit()
    except Exception as e:
        db.session.rollback()
        msg = f"[DB ERROR] {label}: {type(e).__name__}: {str(e)}"
        app.logger.error(msg, exc_info=True)
        print(msg)
        raise

def init_db():
    with app.app_context():
        # 打印当前数据库类型（用于调试，不暴露密码）
        db_uri = app.config.get('SQLALCHEMY_DATABASE_URI', '')
        if 'postgresql' in db_uri:
            print("[INIT] 使用 PostgreSQL 云数据库 ✅")
        else:
            print("[INIT] ⚠️ 警告：使用 SQLite，数据可能在重启后丢失！DATABASE_URL环境变量可能未设置")
        
        db.create_all()
        # 创建管理员账户
        if not User.query.filter_by(username='admin').first():
            admin = User(
                username='admin',
                password_hash=generate_password_hash('admin123456'),
                is_admin=True
            )
            db.session.add(admin)
            safe_commit()
            print("[INIT] 管理员账户已创建：admin / admin123456")
        
        # 创建预设功能分区
        if Section.query.count() == 0:
            sections = [
                {'name': '乡村振兴', 'icon': '🌾', 'sort_order': 1},
                {'name': '党员信息', 'icon': '⭐', 'sort_order': 2},
                {'name': '村民信息', 'icon': '👨‍👩‍👧‍👦', 'sort_order': 3},
                {'name': '养老信息', 'icon': '👴', 'sort_order': 4},
                {'name': '耕地信息', 'icon': '🌱', 'sort_order': 5},
                {'name': '民政信息', 'icon': '📑', 'sort_order': 6},
            ]
            for s in sections:
                db.session.add(Section(**s))
            safe_commit()
            print("[INIT] 6个功能分区已创建")
        
        # 为每个分区创建默认模板
        if Template.query.count() == 0:
            default_templates = {
                '乡村振兴': [
                    {"key": "project_name", "label": "项目名称", "type": "text"},
                    {"key": "implement_date", "label": "实施日期", "type": "date"},
                    {"key": "leader", "label": "负责人", "type": "text"},
                    {"key": "phone", "label": "联系电话", "type": "text"},
                    {"key": "amount", "label": "项目金额(元)", "type": "number"},
                    {"key": "status", "label": "项目状态", "type": "select", "options": ["进行中", "已完成", "暂停"]},
                    {"key": "remark", "label": "备注", "type": "text"}
                ],
                '党员信息': [
                    {"key": "name", "label": "姓名", "type": "text"},
                    {"key": "id_card", "label": "身份证号", "type": "id_card"},
                    {"key": "gender", "label": "性别", "type": "auto", "source": "id_card_gender"},
                    {"key": "age", "label": "年龄", "type": "auto", "source": "id_card_age"},
                    {"key": "join_date", "label": "入党日期", "type": "date"},
                    {"key": "phone", "label": "联系电话", "type": "text"},
                    {"key": "branch", "label": "所在支部", "type": "text"},
                    {"key": "remark", "label": "备注", "type": "text"}
                ],
                '村民信息': [
                    {"key": "name", "label": "姓名", "type": "text"},
                    {"key": "id_card", "label": "身份证号", "type": "id_card"},
                    {"key": "householder_rel", "label": "与户主关系", "type": "select", "options": ["户主", "三女", "三子", "二女", "五子", "儿媳", "养女或继女", "四子", "外孙女", "外孙子", "外甥", "外甥女", "夫", "女", "女婿", "妹妹", "妻", "姐姐", "子", "孙女", "孙媳妇或外孙媳妇", "孙子", "岳母", "曾孙女或曾外孙女", "次子", "母亲", "父亲", "配偶", "长女", "长子", "非亲属"]},
                    {"key": "gender", "label": "性别", "type": "auto", "source": "id_card_gender"},
                    {"key": "age", "label": "年龄", "type": "auto", "source": "id_card_age"},
                    {"key": "phone", "label": "联系电话", "type": "text"},
                    {"key": "address", "label": "家庭住址", "type": "text"},
                    {"key": "remark", "label": "备注", "type": "text"}
                ],
                '养老信息': [
                    {"key": "name", "label": "姓名", "type": "text"},
                    {"key": "id_card", "label": "身份证号", "type": "id_card"},
                    {"key": "gender", "label": "性别", "type": "auto", "source": "id_card_gender"},
                    {"key": "age", "label": "年龄", "type": "auto", "source": "id_card_age"},
                    {"key": "insurance_no", "label": "养老保险号", "type": "text"},
                    {"key": "level", "label": "缴费档次", "type": "select", "options": ["一档", "二档", "三档", "四档", "五档"]},
                    {"key": "phone", "label": "联系电话", "type": "text"},
                    {"key": "remark", "label": "备注", "type": "text"}
                ],
                '耕地信息': [
                    {"key": "name", "label": "户主姓名", "type": "text"},
                    {"key": "id_card", "label": "身份证号", "type": "id_card"},
                    {"key": "gender", "label": "性别", "type": "auto", "source": "id_card_gender"},
                    {"key": "age", "label": "年龄", "type": "auto", "source": "id_card_age"},
                    {"key": "plot_no", "label": "地块编号", "type": "text"},
                    {"key": "area", "label": "耕地面积(亩)", "type": "number"},
                    {"key": "land_type", "label": "土地类型", "type": "select", "options": ["水田", "旱地", "林地", "宅基地", "其他"]},
                    {"key": "phone", "label": "联系电话", "type": "text"},
                    {"key": "remark", "label": "备注", "type": "text"}
                ],
                '资料台账': [
                    {"key": "doc_name", "label": "资料名称", "type": "text"},
                    {"key": "doc_no", "label": "资料编号", "type": "text"},
                    {"key": "archive_date", "label": "归档日期", "type": "date"},
                    {"key": "keeper", "label": "保管人", "type": "text"},
                    {"key": "doc_type", "label": "资料类型", "type": "select", "options": ["纸质", "电子", "音视频", "实物"]},
                    {"key": "remark", "label": "备注", "type": "text"}
                ]
            }
            for section in Section.query.all():
                cols = default_templates.get(section.name, [])
                if cols:
                    t = Template(
                        section_id=section.id,
                        name=f"{section.name}默认模板",
                        columns=json.dumps(cols, ensure_ascii=False)
                    )
                    db.session.add(t)
            safe_commit()
            print("[INIT] 默认模板已创建")

# ===================== 路由：页面 =====================

@app.route('/')
def index():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    sections = Section.query.order_by(Section.sort_order).all()
    home_banner = get_config('home_banner', '')
    return render_template('index.html', sections=sections, user=session.get('username'), home_banner=home_banner)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        user = User.query.filter_by(username=username).first()
        if user and check_password_hash(user.password_hash, password):
            session['user_id'] = user.id
            session['username'] = user.username
            session['is_admin'] = user.is_admin
            return redirect(url_for('index'))
        else:
            flash('用户名或密码错误', 'error')
    return render_template('login.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        confirm = request.form.get('confirm', '').strip()
        if not username or not password:
            flash('用户名和密码不能为空', 'error')
        elif len(password) < 6:
            flash('密码至少6位', 'error')
        elif password != confirm:
            flash('两次输入的密码不一致', 'error')
        elif User.query.filter_by(username=username).first():
            flash('用户名已被注册，请更换', 'error')
        else:
            user = User(
                username=username,
                password_hash=generate_password_hash(password),
                is_admin=False
            )
            db.session.add(user)
            safe_commit()
            flash('注册成功，请登录', 'success')
            return redirect(url_for('login'))
    return render_template('register.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/section/<int:section_id>')
def section_view(section_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    section = Section.query.get_or_404(section_id)
    subsections = SubSection.query.filter_by(section_id=section.id).order_by(SubSection.sort_order).all()
    
    if subsections:
        # 有子功能，显示子功能列表页
        return render_template('section_subs.html', section=section, subsections=subsections)
    
    # 无子功能，保持原有行为
    template = Template.query.filter_by(section_id=section.id).first()
    if not template:
        flash('该分区暂无模板，请联系管理员创建', 'warning')
        return redirect(url_for('index'))
    columns = json.loads(template.columns) if template else []
    return render_template('section.html', target_type='section', target_id=section.id, 
                          target_name=section.name, target_icon=section.icon,
                          section_id=section.id, template_id=template.id, columns=columns)

@app.route('/subsection/<int:subsection_id>')
def subsection_view(subsection_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    sub = SubSection.query.get_or_404(subsection_id)
    section = Section.query.get_or_404(sub.section_id)
    columns = json.loads(sub.columns) if sub.columns else []
    return render_template('section.html', target_type='subsection', target_id=sub.id,
                          target_name=sub.name, target_icon=sub.icon,
                          section_id=section.id, template_id=None, columns=columns)

@app.route('/admin')
def admin():
    if 'user_id' not in session or not session.get('is_admin'):
        flash('需要管理员权限', 'error')
        return redirect(url_for('login'))
    sections = Section.query.order_by(Section.sort_order).all()
    templates = Template.query.all()
    for t in templates:
        t.columns_list = json.loads(t.columns)
    users = User.query.all()
    home_banner = get_config('home_banner', '')
    return render_template('admin.html', sections=sections, templates=templates, users=users, home_banner=home_banner)

# ===================== 路由：数据API =====================

@app.route('/api/records/<int:section_id>')
def api_records(section_id):
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    template = Template.query.filter_by(section_id=section_id).first()
    if not template:
        return jsonify({'records': []})
    records = Record.query.filter_by(section_id=section_id, template_id=template.id).all()
    data = []
    for r in records:
        item = json.loads(r.data)
        item['_id'] = r.id
        data.append(item)
    return jsonify({'records': data, 'template_id': template.id})

@app.route('/api/records/save', methods=['POST'])
def api_save_records():
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    payload = request.get_json()
    section_id = payload.get('section_id')
    template_id = payload.get('template_id')
    rows = payload.get('rows', [])
    
    template = Template.query.get(template_id)
    columns = json.loads(template.columns) if template else []
    
    # 先删除旧记录（全量替换模式，简化实现）
    Record.query.filter_by(section_id=section_id, template_id=template_id).delete(synchronize_session=False)
    
    saved = 0
    for row in rows:
        if not row or all(v == '' for v in row.values()):
            continue
        # 校验身份证
        for col in columns:
            if col.get('type') == 'id_card':
                val = row.get(col['key'], '')
                if val and not validate_id_card(val):
                    return jsonify({'error': f"身份证校验失败: {val}"}), 400
        # 自动填充
        for col in columns:
            if col.get('type') == 'auto':
                row[col['key']] = auto_fill_value(col, row)
        
        rec = Record(section_id=section_id, template_id=template_id, data=json.dumps(row, ensure_ascii=False))
        db.session.add(rec)
        saved += 1
    safe_commit()
    return jsonify({'success': True, 'saved': saved})

@app.route('/api/records/delete/<int:record_id>', methods=['POST'])
def api_delete_record(record_id):
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    rec = Record.query.get_or_404(record_id)
    # 移入回收站
    del_rec = DeletedRecord(
        original_id=rec.id,
        section_id=rec.section_id,
        template_id=rec.template_id,
        subsection_id=rec.subsection_id,
        data=rec.data,
        deleted_by=session.get('username', 'unknown')
    )
    db.session.add(del_rec)
    db.session.delete(rec)
    safe_commit()
    return jsonify({'success': True, 'deleted_id': del_rec.id})

@app.route('/api/records/batch_delete', methods=['POST'])
def api_batch_delete_records():
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    payload = request.get_json()
    ids = payload.get('ids', [])
    deleted_ids = []
    for rid in ids:
        rec = Record.query.get(rid)
        if rec:
            del_rec = DeletedRecord(
                original_id=rec.id,
                section_id=rec.section_id,
                template_id=rec.template_id,
                subsection_id=rec.subsection_id,
                data=rec.data,
                deleted_by=session.get('username', 'unknown')
            )
            db.session.add(del_rec)
            db.session.delete(rec)
            deleted_ids.append(del_rec.id)
    safe_commit()
    return jsonify({'success': True, 'deleted_count': len(deleted_ids)})

# 回收站 API
@app.route('/api/recycle/<target_type>/<int:target_id>')
def api_recycle_bin(target_type, target_id):
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    query = DeletedRecord.query
    if target_type == 'section':
        query = query.filter_by(section_id=target_id, subsection_id=None)
    elif target_type == 'subsection':
        query = query.filter_by(subsection_id=target_id)
    else:
        return jsonify({'error': '类型错误'}), 400
    records = query.order_by(DeletedRecord.deleted_at.desc()).all()
    return jsonify({'success': True, 'records': [{
        'id': r.id, 'original_id': r.original_id, 'data': json.loads(r.data),
        'deleted_by': r.deleted_by, 'deleted_at': r.deleted_at.strftime('%Y-%m-%d %H:%M') if r.deleted_at else '-'
    } for r in records]})

@app.route('/api/recycle/restore/<int:deleted_id>', methods=['POST'])
def api_restore_record(deleted_id):
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    del_rec = DeletedRecord.query.get_or_404(deleted_id)
    rec = Record(
        section_id=del_rec.section_id,
        template_id=del_rec.template_id,
        subsection_id=del_rec.subsection_id,
        data=del_rec.data
    )
    db.session.add(rec)
    db.session.delete(del_rec)
    safe_commit()
    return jsonify({'success': True, 'restored_id': rec.id})

@app.route('/api/recycle/clear', methods=['POST'])
def api_clear_recycle():
    if 'user_id' not in session or not session.get('is_admin'):
        return jsonify({'error': '无权限'}), 403
    payload = request.get_json()
    target_type = payload.get('target_type')
    target_id = payload.get('target_id')
    query = DeletedRecord.query
    if target_type == 'section':
        query = query.filter_by(section_id=target_id, subsection_id=None)
    elif target_type == 'subsection':
        query = query.filter_by(subsection_id=target_id)
    count = query.count()
    query.delete(synchronize_session=False)
    safe_commit()
    return jsonify({'success': True, 'cleared_count': count})

# 搜索 API
@app.route('/api/search/<target_type>/<int:target_id>')
def api_search_records(target_type, target_id):
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    keyword = request.args.get('q', '').strip()
    age_range = request.args.get('age', '')
    gender = request.args.get('gender', '')
    
    if target_type == 'section':
        records = Record.query.filter_by(section_id=target_id).all()
    elif target_type == 'subsection':
        records = Record.query.filter_by(subsection_id=target_id).all()
    else:
        return jsonify({'error': '类型错误'}), 400
    
    results = []
    for r in records:
        data = json.loads(r.data)
        data['_id'] = r.id
        # 姓名关键词匹配
        name = data.get('name', '')
        id_card = data.get('id_card', '')
        rec_age = data.get('age', '')
        rec_gender = data.get('gender', '')
        
        match = True
        if keyword and keyword not in name and keyword not in id_card:
            match = False
        if gender and rec_gender != gender:
            match = False
        if age_range:
            try:
                age_int = int(rec_age) if rec_age else -1
                if age_range == '0-17' and not (0 <= age_int <= 17):
                    match = False
                elif age_range == '18-35' and not (18 <= age_int <= 35):
                    match = False
                elif age_range == '36-59' and not (36 <= age_int <= 59):
                    match = False
                elif age_range == '60+' and not (age_int >= 60):
                    match = False
            except:
                match = False
        if match:
            results.append(data)
    return jsonify({'success': True, 'records': results, 'total': len(results)})

# ===================== 路由：Excel导入导出 =====================

@app.route('/api/export/<int:section_id>')
def api_export(section_id):
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    section = Section.query.get_or_404(section_id)
    template = Template.query.filter_by(section_id=section_id).first()
    if not template:
        return jsonify({'error': '无模板'}), 400
    columns = json.loads(template.columns)
    records = Record.query.filter_by(section_id=section_id, template_id=template.id).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = section.name
    ws.append([c['label'] for c in columns])
    for r in records:
        row_data = json.loads(r.data)
        ws.append([row_data.get(c['key'], '') for c in columns])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"{section.name}_{datetime.datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(output, download_name=filename, as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/api/import/<int:section_id>', methods=['POST'])
def api_import(section_id):
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    template = Template.query.filter_by(section_id=section_id).first()
    if not template:
        return jsonify({'error': '无模板'}), 400
    columns = json.loads(template.columns)
    
    file = request.files.get('file')
    if not file:
        return jsonify({'error': '未上传文件'}), 400
    try:
        wb = load_workbook(file)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        # 映射列
        col_map = {}
        for idx, h in enumerate(headers):
            for c in columns:
                if c['label'] == h:
                    col_map[c['key']] = idx
                    break
        
        rows = []
        errors = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            item = {}
            for c in columns:
                idx = col_map.get(c['key'])
                if idx is not None and idx < len(row):
                    val = row[idx]
                    item[c['key']] = str(val) if val is not None else ''
                else:
                    item[c['key']] = ''
            # 校验身份证
            for c in columns:
                if c.get('type') == 'id_card':
                    val = item.get(c['key'], '')
                    if val and not validate_id_card(val):
                        errors.append({'row': row_idx, 'id_card': val, 'reason': '身份证校验不通过'})
            # 自动填充
            for c in columns:
                if c.get('type') == 'auto':
                    item[c['key']] = auto_fill_value(c, item)
            if not all(v == '' for v in item.values()):
                rows.append(item)
        
        if errors:
            return jsonify({'error': f'导入失败，共发现 {len(errors)} 条身份证错误', 'errors': errors}), 400
        
        # 保存
        Record.query.filter_by(section_id=section_id, template_id=template.id).delete(synchronize_session=False)
        for item in rows:
            rec = Record(section_id=section_id, template_id=template.id, data=json.dumps(item, ensure_ascii=False))
            db.session.add(rec)
        safe_commit()
        return jsonify({'success': True, 'imported': len(rows)})
    except Exception as e:
        return jsonify({'error': f'导入失败: {str(e)}'}), 400

# ===================== 路由：管理API（分区） =====================

@app.route('/api/admin/sections', methods=['GET', 'POST'])
def api_admin_sections():
    if 'user_id' not in session or not session.get('is_admin'):
        return jsonify({'error': '无权限'}), 403
    if request.method == 'GET':
        sections = Section.query.order_by(Section.sort_order).all()
        return jsonify([{'id': s.id, 'name': s.name, 'icon': s.icon, 'sort_order': s.sort_order} for s in sections])
    elif request.method == 'POST':
        data = request.get_json()
        s = Section(
            name=data.get('name', ''),
            icon=data.get('icon', '📁'),
            sort_order=data.get('sort_order', 0),
            description=data.get('description', ''),
            bg_color=data.get('bg_color', ''),
            text_color=data.get('text_color', ''),
            accent_color=data.get('accent_color', '')
        )
        db.session.add(s)
        safe_commit()
        # 自动创建默认模板
        default_cols = [
            {"key": "name", "label": "名称", "type": "text"},
            {"key": "id_card", "label": "身份证号", "type": "id_card"},
            {"key": "gender", "label": "性别", "type": "auto", "source": "id_card_gender"},
            {"key": "age", "label": "年龄", "type": "auto", "source": "id_card_age"},
            {"key": "phone", "label": "联系电话", "type": "text"},
            {"key": "remark", "label": "备注", "type": "text"}
        ]
        t = Template(section_id=s.id, name=f"{s.name}默认模板", columns=json.dumps(default_cols, ensure_ascii=False))
        db.session.add(t)
        safe_commit()
        return jsonify({'success': True, 'id': s.id})

@app.route('/api/admin/sections/<int:section_id>', methods=['PUT', 'DELETE'])
def api_admin_section(section_id):
    if 'user_id' not in session or not session.get('is_admin'):
        return jsonify({'error': '无权限'}), 403
    s = Section.query.get_or_404(section_id)
    if request.method == 'PUT':
        data = request.get_json()
        s.name = data.get('name', s.name)
        s.icon = data.get('icon', s.icon)
        s.sort_order = data.get('sort_order', s.sort_order)
        s.bg_color = data.get('bg_color', s.bg_color)
        s.text_color = data.get('text_color', s.text_color)
        s.accent_color = data.get('accent_color', s.accent_color)
        s.description = data.get('description', s.description)
        safe_commit()
        return jsonify({'success': True})
    elif request.method == 'DELETE':
        Record.query.filter_by(section_id=section_id).delete(synchronize_session=False)
        Template.query.filter_by(section_id=section_id).delete(synchronize_session=False)
        SubSection.query.filter_by(section_id=section_id).delete(synchronize_session=False)
        db.session.delete(s)
        safe_commit()
        return jsonify({'success': True})

# ===================== 路由：管理API（模板） =====================

@app.route('/api/admin/templates', methods=['GET', 'POST'])
def api_admin_templates():
    if 'user_id' not in session or not session.get('is_admin'):
        return jsonify({'error': '无权限'}), 403
    if request.method == 'GET':
        templates = Template.query.all()
        return jsonify([{
            'id': t.id, 'section_id': t.section_id, 'name': t.name,
            'columns': json.loads(t.columns), 'section_name': t.section.name if t.section else ''
        } for t in templates])
    elif request.method == 'POST':
        data = request.get_json()
        t = Template(
            section_id=data.get('section_id'),
            name=data.get('name', ''),
            columns=json.dumps(data.get('columns', []), ensure_ascii=False)
        )
        db.session.add(t)
        safe_commit()
        return jsonify({'success': True, 'id': t.id})

@app.route('/api/admin/templates/<int:template_id>', methods=['PUT', 'DELETE'])
def api_admin_template(template_id):
    if 'user_id' not in session or not session.get('is_admin'):
        return jsonify({'error': '无权限'}), 403
    t = Template.query.get_or_404(template_id)
    if request.method == 'PUT':
        data = request.get_json()
        t.name = data.get('name', t.name)
        t.columns = json.dumps(data.get('columns', []), ensure_ascii=False)
        safe_commit()
        return jsonify({'success': True})
    elif request.method == 'DELETE':
        Record.query.filter_by(template_id=template_id).delete(synchronize_session=False)
        db.session.delete(t)
        safe_commit()
        return jsonify({'success': True})

# ===================== 路由：文件上传（图片） =====================

def save_image_to_db(filename, file_obj, mime_type='image/png'):
    """将上传的图片以base64存入数据库，返回URL路径"""
    file_obj.seek(0)
    data = file_obj.read()
    data_b64 = base64.b64encode(data).decode('utf-8')
    # 删除同名旧记录
    old = UploadedImage.query.filter_by(filename=filename).first()
    if old:
        old.data_b64 = data_b64
        old.mime_type = mime_type
    else:
        img = UploadedImage(filename=filename, mime_type=mime_type, data_b64=data_b64)
        db.session.add(img)
    safe_commit()
    return f'/uploads/{filename}'

def delete_image_from_db(filename):
    """从数据库删除图片"""
    img = UploadedImage.query.filter_by(filename=filename).first()
    if img:
        db.session.delete(img)
        safe_commit()
        return True
    return False

def get_ext_mime(ext):
    """根据扩展名返回mime类型"""
    mime_map = {
        'png': 'image/png', 'jpg': 'image/jpeg', 'jpeg': 'image/jpeg',
        'gif': 'image/gif', 'webp': 'image/webp', 'bmp': 'image/bmp'
    }
    return mime_map.get(ext.lower(), 'image/png')

@app.route('/uploads/<filename>')
def serve_uploaded_image(filename):
    """从数据库读取并返回上传的图片"""
    img = UploadedImage.query.filter_by(filename=filename).first()
    if img and img.data_b64:
        data = base64.b64decode(img.data_b64)
        return Response(data, mimetype=img.mime_type)
    # 本地开发模式：回退到文件系统
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    if os.path.exists(filepath):
        return send_file(filepath)
    return 'Not Found', 404

@app.route('/api/upload/banner', methods=['POST'])
def upload_banner():
    if 'user_id' not in session or not session.get('is_admin'):
        return jsonify({'error': '无权限'}), 403
    file = request.files.get('file')
    if not file or not allowed_file(file.filename):
        return jsonify({'error': '请上传图片文件（png/jpg/gif/webp/bmp）'}), 400
    ext = file.filename.rsplit('.', 1)[1].lower()
    filename = f'home_banner.{ext}'
    # 删除旧图片（所有扩展名）
    for old_ext in ['png', 'jpg', 'jpeg', 'gif', 'webp', 'bmp']:
        if old_ext != ext:
            delete_image_from_db(f'home_banner.{old_ext}')
    url = save_image_to_db(filename, file, get_ext_mime(ext))
    set_config('home_banner', filename)
    return jsonify({'success': True, 'filename': filename, 'url': url})

@app.route('/api/upload/section_icon/<int:section_id>', methods=['POST'])
def upload_section_icon(section_id):
    if 'user_id' not in session or not session.get('is_admin'):
        return jsonify({'error': '无权限'}), 403
    section = Section.query.get_or_404(section_id)
    file = request.files.get('file')
    if not file or not allowed_file(file.filename):
        return jsonify({'error': '请上传图片文件'}), 400
    ext = file.filename.rsplit('.', 1)[1].lower()
    filename = f'section_icon_{section_id}.{ext}'
    # 删除旧图标（所有扩展名）
    for old_ext in ['png', 'jpg', 'jpeg', 'gif', 'webp', 'bmp']:
        if old_ext != ext:
            delete_image_from_db(f'section_icon_{section_id}.{old_ext}')
    url = save_image_to_db(filename, file, get_ext_mime(ext))
    section.icon_image = filename
    safe_commit()
    return jsonify({'success': True, 'filename': filename, 'url': url})

@app.route('/api/upload/delete_banner', methods=['POST'])
def delete_banner():
    if 'user_id' not in session or not session.get('is_admin'):
        return jsonify({'error': '无权限'}), 403
    filename = get_config('home_banner', '')
    if filename:
        delete_image_from_db(filename)
        set_config('home_banner', '')
    return jsonify({'success': True})

@app.route('/api/upload/delete_section_icon/<int:section_id>', methods=['POST'])
def delete_section_icon(section_id):
    if 'user_id' not in session or not session.get('is_admin'):
        return jsonify({'error': '无权限'}), 403
    section = Section.query.get_or_404(section_id)
    if section.icon_image:
        delete_image_from_db(section.icon_image)
        section.icon_image = ''
        safe_commit()
    return jsonify({'success': True})

# ===================== 路由：管理API（账户） =====================

@app.route('/api/admin/users/password', methods=['POST'])
def api_admin_change_password():
    if 'user_id' not in session or not session.get('is_admin'):
        return jsonify({'error': '无权限'}), 403
    data = request.get_json()
    old_password = data.get('old_password', '')
    new_password = data.get('new_password', '')
    if not new_password or len(new_password) < 6:
        return jsonify({'error': '新密码至少6位'}), 400
    user = User.query.get(session['user_id'])
    if not check_password_hash(user.password_hash, old_password):
        return jsonify({'error': '原密码错误'}), 400
    user.password_hash = generate_password_hash(new_password)
    safe_commit()
    return jsonify({'success': True})

@app.route('/api/admin/users', methods=['POST'])
def api_admin_add_user():
    if 'user_id' not in session or not session.get('is_admin'):
        return jsonify({'error': '无权限'}), 403
    data = request.get_json()
    username = data.get('username', '').strip()
    password = data.get('password', '').strip()
    if not username or not password:
        return jsonify({'error': '用户名和密码不能为空'}), 400
    if User.query.filter_by(username=username).first():
        return jsonify({'error': '用户名已存在'}), 400
    u = User(username=username, password_hash=generate_password_hash(password), is_admin=False)
    db.session.add(u)
    safe_commit()
    return jsonify({'success': True, 'id': u.id})

@app.route('/api/admin/users/<int:user_id>', methods=['DELETE'])
def api_admin_delete_user(user_id):
    if 'user_id' not in session or not session.get('is_admin'):
        return jsonify({'error': '无权限'}), 403
    u = User.query.get_or_404(user_id)
    if u.username == 'admin':
        return jsonify({'error': '不能删除默认管理员'}), 400
    db.session.delete(u)
    safe_commit()
    return jsonify({'success': True})

# ===================== 路由：子功能数据API =====================

@app.route('/api/subsections/<int:section_id>')
def api_subsections(section_id):
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    subs = SubSection.query.filter_by(section_id=section_id).order_by(SubSection.sort_order).all()
    return jsonify([{
        'id': s.id, 'name': s.name, 'icon': s.icon, 'icon_image': s.icon_image,
        'description': s.description, 'sort_order': s.sort_order,
        'section_id': s.section_id
    } for s in subs])

@app.route('/api/subsections/records/<int:subsection_id>')
def api_subsection_records(subsection_id):
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    sub = SubSection.query.get_or_404(subsection_id)
    records = Record.query.filter_by(subsection_id=subsection_id).all()
    data = []
    for r in records:
        item = json.loads(r.data)
        item['_id'] = r.id
        data.append(item)
    return jsonify({'records': data, 'subsection_id': subsection_id})

@app.route('/api/subsections/records/save', methods=['POST'])
def api_save_subsection_records():
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    payload = request.get_json()
    subsection_id = payload.get('subsection_id')
    rows = payload.get('rows', [])
    
    sub = SubSection.query.get(subsection_id)
    if not sub:
        return jsonify({'error': '子功能不存在'}), 404
    columns = json.loads(sub.columns) if sub.columns else []
    
    Record.query.filter_by(subsection_id=subsection_id).delete(synchronize_session=False)
    
    saved = 0
    for row in rows:
        if not row or all(v == '' for v in row.values()):
            continue
        for col in columns:
            if col.get('type') == 'id_card':
                val = row.get(col['key'], '')
                if val and not validate_id_card(val):
                    return jsonify({'error': f"身份证校验失败: {val}"}), 400
        for col in columns:
            if col.get('type') == 'auto':
                row[col['key']] = auto_fill_value(col, row)
        
        rec = Record(section_id=sub.section_id, template_id=0, subsection_id=subsection_id, 
                     data=json.dumps(row, ensure_ascii=False))
        db.session.add(rec)
        saved += 1
    safe_commit()
    return jsonify({'success': True, 'saved': saved})

@app.route('/api/subsections/export/<int:subsection_id>')
def api_subsection_export(subsection_id):
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    sub = SubSection.query.get_or_404(subsection_id)
    columns = json.loads(sub.columns) if sub.columns else []
    records = Record.query.filter_by(subsection_id=subsection_id).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = sub.name
    ws.append([c['label'] for c in columns])
    for r in records:
        row_data = json.loads(r.data)
        ws.append([row_data.get(c['key'], '') for c in columns])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"{sub.name}_{datetime.datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(output, download_name=filename, as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/api/stats/subsection/<int:subsection_id>')
def api_subsection_stats(subsection_id):
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    sub = SubSection.query.get_or_404(subsection_id)
    records = Record.query.filter_by(subsection_id=subsection_id).all()
    return jsonify(_build_stats(records, sub.name))


def _build_stats(records, name):
    """构建统计数据的通用函数"""
    total = 0
    gender_stats = {'男': 0, '女': 0, '未知': 0}
    age_groups = {'0-17岁': 0, '18-35岁': 0, '36-59岁': 0, '60岁以上': 0, '未知': 0}
    household_dist = {}
    householder_count = 0

    for r in records:
        data = json.loads(r.data)
        total += 1
        # gender
        g = data.get('gender', '').strip()
        if g in gender_stats:
            gender_stats[g] += 1
        else:
            gender_stats['未知'] += 1
        # age
        age_str = data.get('age', '').strip()
        try:
            age = int(age_str)
            if age < 18:
                age_groups['0-17岁'] += 1
            elif age <= 35:
                age_groups['18-35岁'] += 1
            elif age <= 59:
                age_groups['36-59岁'] += 1
            else:
                age_groups['60岁以上'] += 1
        except (ValueError, TypeError):
            age_groups['未知'] += 1
        # 与户主关系
        rel = data.get('householder_rel', '').strip()
        if rel:
            household_dist[rel] = household_dist.get(rel, 0) + 1
            if rel == '户主':
                householder_count += 1

    return {
        'success': True,
        'total': total,
        'householder_count': householder_count,
        'household_dist': household_dist,
        'gender': gender_stats,
        'age': age_groups,
        'name': name
    }

@app.route('/api/stats/section/<int:section_id>')
def api_section_stats(section_id):
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    section = Section.query.get_or_404(section_id)
    # 统计该主分区下所有记录（包括子功能和直接模板）
    records = Record.query.filter_by(section_id=section_id).all()
    return jsonify(_build_stats(records, section.name))

@app.route('/api/subsections/import/<int:subsection_id>', methods=['POST'])
def api_subsection_import(subsection_id):
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    sub = SubSection.query.get_or_404(subsection_id)
    columns = json.loads(sub.columns) if sub.columns else []
    
    file = request.files.get('file')
    if not file:
        return jsonify({'error': '未上传文件'}), 400
    try:
        wb = load_workbook(file)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        col_map = {}
        for idx, h in enumerate(headers):
            for c in columns:
                if c['label'] == h:
                    col_map[c['key']] = idx
                    break
        
        rows = []
        errors = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            item = {}
            for c in columns:
                idx = col_map.get(c['key'])
                if idx is not None and idx < len(row):
                    val = row[idx]
                    item[c['key']] = str(val) if val is not None else ''
                else:
                    item[c['key']] = ''
            for c in columns:
                if c.get('type') == 'id_card':
                    val = item.get(c['key'], '')
                    if val and not validate_id_card(val):
                        errors.append({'row': row_idx, 'id_card': val, 'reason': '身份证校验不通过'})
            for c in columns:
                if c.get('type') == 'auto':
                    item[c['key']] = auto_fill_value(c, item)
            if not all(v == '' for v in item.values()):
                rows.append(item)
        
        if errors:
            return jsonify({'error': f'导入失败，共发现 {len(errors)} 条身份证错误', 'errors': errors}), 400
        
        Record.query.filter_by(subsection_id=subsection_id).delete(synchronize_session=False)
        for item in rows:
            rec = Record(section_id=sub.section_id, template_id=0, subsection_id=subsection_id, 
                         data=json.dumps(item, ensure_ascii=False))
            db.session.add(rec)
        safe_commit()
        return jsonify({'success': True, 'imported': len(rows)})
    except Exception as e:
        return jsonify({'error': f'导入失败: {str(e)}'}), 400

# ===================== 路由：管理API（子功能） =====================

@app.route('/api/admin/subsections', methods=['GET', 'POST'])
def api_admin_subsections():
    if 'user_id' not in session or not session.get('is_admin'):
        return jsonify({'error': '无权限'}), 403
    if request.method == 'GET':
        section_id = request.args.get('section_id', type=int)
        query = SubSection.query
        if section_id:
            query = query.filter_by(section_id=section_id)
        subs = query.order_by(SubSection.sort_order).all()
        return jsonify([{
            'id': s.id, 'section_id': s.section_id, 'name': s.name, 'icon': s.icon,
            'icon_image': s.icon_image, 'description': s.description,
            'sort_order': s.sort_order, 'columns': json.loads(s.columns),
            'section_name': s.section.name if s.section else '未知'
        } for s in subs])
    elif request.method == 'POST':
        data = request.get_json()
        sub = SubSection(
            section_id=data.get('section_id'),
            name=data.get('name', ''),
            icon=data.get('icon', '📄'),
            description=data.get('description', ''),
            sort_order=data.get('sort_order', 0),
            columns=json.dumps(data.get('columns', []), ensure_ascii=False)
        )
        db.session.add(sub)
        safe_commit()
        return jsonify({'success': True, 'id': sub.id})

@app.route('/api/admin/subsections/<int:subsection_id>', methods=['PUT', 'DELETE'])
def api_admin_subsection(subsection_id):
    if 'user_id' not in session or not session.get('is_admin'):
        return jsonify({'error': '无权限'}), 403
    sub = SubSection.query.get_or_404(subsection_id)
    if request.method == 'PUT':
        data = request.get_json()
        sub.name = data.get('name', sub.name)
        sub.icon = data.get('icon', sub.icon)
        sub.description = data.get('description', sub.description)
        sub.sort_order = data.get('sort_order', sub.sort_order)
        if 'columns' in data:
            sub.columns = json.dumps(data.get('columns', []), ensure_ascii=False)
        safe_commit()
        return jsonify({'success': True})
    elif request.method == 'DELETE':
        Record.query.filter_by(subsection_id=subsection_id).delete(synchronize_session=False)
        db.session.delete(sub)
        safe_commit()
        return jsonify({'success': True})

@app.route('/api/upload/subsection_icon/<int:subsection_id>', methods=['POST'])
def upload_subsection_icon(subsection_id):
    if 'user_id' not in session or not session.get('is_admin'):
        return jsonify({'error': '无权限'}), 403
    sub = SubSection.query.get_or_404(subsection_id)
    file = request.files.get('file')
    if not file or not allowed_file(file.filename):
        return jsonify({'error': '请上传图片文件'}), 400
    ext = file.filename.rsplit('.', 1)[1].lower()
    filename = f'subsection_icon_{subsection_id}.{ext}'
    for old_ext in ['png', 'jpg', 'jpeg', 'gif', 'webp', 'bmp']:
        if old_ext != ext:
            delete_image_from_db(f'subsection_icon_{subsection_id}.{old_ext}')
    url = save_image_to_db(filename, file, get_ext_mime(ext))
    sub.icon_image = filename
    safe_commit()
    return jsonify({'success': True, 'filename': filename, 'url': url})

@app.route('/api/upload/delete_subsection_icon/<int:subsection_id>', methods=['POST'])
def delete_subsection_icon(subsection_id):
    if 'user_id' not in session or not session.get('is_admin'):
        return jsonify({'error': '无权限'}), 403
    sub = SubSection.query.get_or_404(subsection_id)
    if sub.icon_image:
        delete_image_from_db(sub.icon_image)
        sub.icon_image = ''
        safe_commit()
    return jsonify({'success': True})

# ===================== 启动入口 =====================

# 生产环境（gunicorn）导入时自动初始化数据库
init_db()

if __name__ == '__main__':
    # 本地开发模式
    app.run(host='0.0.0.0', port=int(os.environ.get('PORT', 5000)), debug=True)
